import os
import io
import re
import zipfile
import logging
from datetime import datetime
from typing import Dict, Tuple, Optional
from functools import wraps
import time

from flask import Flask, request, send_file, jsonify, Response
from werkzeug.exceptions import RequestEntityTooLarge
from openpyxl import load_workbook, Workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)

# Configuration
MAX_FILE_SIZE_MB = int(os.environ.get('MAX_FILE_SIZE_MB', 50))  # Size in MB (default 50MB)
MAX_FILE_SIZE = MAX_FILE_SIZE_MB * 1024 * 1024  # Convert MB to bytes
MAX_SHEETS = int(os.environ.get('MAX_SHEETS', 100))  # Maximum number of sheets to process
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm', 'xlsb'}
PORT = int(os.environ.get('PORT', 3070))

app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE


def timing_decorator(f):
    """Decorator to measure and log function execution time."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = f(*args, **kwargs)
        duration = time.time() - start
        logger.info(f"{f.__name__} took {duration:.2f} seconds")
        return result
    return wrapper


def sanitize_filename(filename: str) -> str:
    """
    Sanitize filename for safe file system usage.
    
    Args:
        filename: Original filename
        
    Returns:
        Sanitized filename safe for file systems
    """
    # Remove invalid characters for filenames
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # Remove leading/trailing spaces and dots
    sanitized = sanitized.strip('. ')
    # Limit length to prevent filesystem issues
    max_length = 100
    if len(sanitized) > max_length:
        name, ext = os.path.splitext(sanitized)
        sanitized = name[:max_length - len(ext)] + ext
    # Ensure non-empty filename
    if not sanitized:
        sanitized = 'unnamed'
    return sanitized


def allowed_file(filename: str) -> bool:
    """
    Check if file has an allowed extension.
    
    Args:
        filename: Name of the file to check
        
    Returns:
        True if file extension is allowed, False otherwise
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_excel_file(file_bytes: bytes) -> Tuple[bool, Optional[str]]:
    """
    Validate that the bytes represent a valid Excel file.
    
    Args:
        file_bytes: Raw bytes of the file
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    try:
        with io.BytesIO(file_bytes) as buffer:
            load_workbook(buffer, read_only=True, data_only=False)
        return True, None
    except Exception as e:
        logger.warning(f"Excel validation failed: {str(e)}")
        return False, f"Invalid Excel file: {str(e)}"


@timing_decorator
def split_excel_by_sheets_simple(
    file_bytes: bytes, 
    original_filename: str
) -> Tuple[Dict[str, bytes], Optional[str]]:
    """
    Split an Excel file into separate files for each sheet.
    Preserves data, formulas, merged cells, and basic structure.
    
    Args:
        file_bytes: Raw bytes of the Excel file
        original_filename: Original name of the uploaded file
        
    Returns:
        Tuple of (dict mapping sheet names to bytes, error message)
    """
    split_files = {}
    base_name = os.path.splitext(original_filename)[0]
    
    try:
        # Load the workbook
        with io.BytesIO(file_bytes) as input_buffer:
            source_wb = load_workbook(
                input_buffer, 
                read_only=False, 
                keep_vba=False,
                data_only=False,
                keep_links=True
            )
            
            # Check sheet count
            if len(source_wb.sheetnames) > MAX_SHEETS:
                return {}, f"File contains too many sheets ({len(source_wb.sheetnames)}). Maximum allowed: {MAX_SHEETS}"
            
            # Filter out hidden sheets
            visible_sheets = []
            hidden_sheets = []
            for sheet_name in source_wb.sheetnames:
                sheet = source_wb[sheet_name]
                if sheet.sheet_state == 'visible':
                    visible_sheets.append(sheet_name)
                else:
                    hidden_sheets.append(sheet_name)
            
            logger.info(f"Processing {len(visible_sheets)} visible sheets from {original_filename}")
            if hidden_sheets:
                logger.info(f"Skipping {len(hidden_sheets)} hidden sheets: {', '.join(hidden_sheets)}")
            
            # Process each visible sheet
            for sheet_name in visible_sheets:
                try:
                    logger.info(f"Processing sheet: {sheet_name}")
                    source_ws = source_wb[sheet_name]
                    
                    # Create new workbook for this sheet
                    new_wb = Workbook()
                    
                    # Remove default sheet and create new one with same name
                    new_wb.remove(new_wb.active)
                    target_ws = new_wb.create_sheet(title=sheet_name)
                    
                    # Copy column widths
                    for col_letter, col_dim in source_ws.column_dimensions.items():
                        if col_dim.width:
                            target_ws.column_dimensions[col_letter].width = col_dim.width
                    
                    # Copy row heights
                    for row_num, row_dim in source_ws.row_dimensions.items():
                        if row_dim.height:
                            target_ws.row_dimensions[row_num].height = row_dim.height
                    
                    # Copy all cells with values and basic formatting
                    for row in source_ws.iter_rows():
                        for cell in row:
                            target_cell = target_ws.cell(row=cell.row, column=cell.column)
                            
                            # Copy value
                            target_cell.value = cell.value
                            
                            # Copy basic formatting
                            if cell.has_style:
                                try:
                                    target_cell.font = cell.font.copy()
                                    target_cell.fill = cell.fill.copy()
                                    target_cell.border = cell.border.copy()
                                    target_cell.alignment = cell.alignment.copy()
                                    target_cell.number_format = cell.number_format
                                except Exception as e:
                                    logger.debug(f"Could not copy cell style: {e}")
                    
                    # Copy merged cells
                    for merged_range in source_ws.merged_cells.ranges:
                        target_ws.merge_cells(str(merged_range))
                    
                    # Save to bytes
                    output_buffer = io.BytesIO()
                    new_wb.save(output_buffer)
                    output_buffer.seek(0)
                    
                    # Create filename for this sheet
                    sanitized_sheet_name = sanitize_filename(sheet_name)
                    output_filename = f"{base_name}_{sanitized_sheet_name}.xlsx"
                    
                    split_files[output_filename] = output_buffer.getvalue()
                    logger.info(f"Successfully processed sheet: {sheet_name}")
                    
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
                    # Continue processing other sheets even if one fails
                    continue
        
        if not split_files:
            return {}, "No sheets could be processed successfully"
            
        return split_files, None
        
    except Exception as e:
        logger.error(f"Error splitting Excel file: {str(e)}")
        return {}, f"Error processing Excel file: {str(e)}"


def create_zip_response(files_dict: Dict[str, bytes], base_filename: str) -> Response:
    """
    Create a ZIP file response containing multiple files.
    
    Args:
        files_dict: Dictionary mapping filenames to file bytes
        base_filename: Base name for the ZIP file
        
    Returns:
        Flask Response object with ZIP file
    """
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, file_bytes in files_dict.items():
            zip_file.writestr(filename, file_bytes)
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"{base_filename}_split.zip"
    )


@app.route('/', methods=['GET'])
def index():
    """Service information endpoint."""
    return jsonify({
        'service': 'Excel File Splitter',
        'version': '2.1.0',
        'description': 'Split Excel files by sheets while preserving data and structure',
        'endpoints': {
            'POST /split-excel': 'Upload Excel file to split',
            'GET /health': 'Health check endpoint',
            'GET /': 'Service information'
        },
        'features': [
            'Preserves cell values and formulas',
            'Maintains merged cells',
            'Keeps column widths and row heights',
            'Preserves basic cell formatting',
            'Maintains number formats'
        ],
        'configuration': {
            'max_file_size': f"{MAX_FILE_SIZE / (1024*1024):.1f} MB",
            'max_sheets': MAX_SHEETS,
            'allowed_extensions': list(ALLOWED_EXTENSIONS)
        },
        'timestamp': datetime.utcnow().isoformat()
    })


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint for monitoring."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'service': 'excel-splitter'
    }), 200


@app.route('/split-excel', methods=['POST'])
def split_excel():
    """
    Main endpoint to split Excel files by sheets.
    
    Accepts: multipart/form-data with 'file' field
    Returns: 
        - Single sheet: Direct Excel file
        - Multiple sheets: ZIP file containing all sheets
    """
    try:
        # Validate request has file
        if 'file' not in request.files:
            logger.warning("No file in request")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        # Validate file selection
        if file.filename == '':
            logger.warning("Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not allowed_file(file.filename):
            logger.warning(f"Invalid file extension: {file.filename}")
            return jsonify({
                'error': f'Invalid file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}'
            }), 400
        
        # Read file bytes
        file_bytes = file.read()
        
        # Log file info
        file_size_mb = len(file_bytes) / (1024 * 1024)
        logger.info(f"Received file: {file.filename} ({file_size_mb:.2f} MB)")
        
        # Validate Excel file structure
        is_valid, error_msg = validate_excel_file(file_bytes)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        # Split the Excel file
        split_files, error = split_excel_by_sheets_simple(
            file_bytes, 
            file.filename
        )
        
        if error:
            logger.error(f"Splitting failed: {error}")
            return jsonify({'error': error}), 500
        
        if not split_files:
            return jsonify({'error': 'No sheets found in Excel file'}), 400
        
        # Prepare response based on number of sheets
        base_name = os.path.splitext(file.filename)[0]
        
        if len(split_files) == 1:
            # Single sheet - return Excel file directly
            filename, file_bytes = next(iter(split_files.items()))
            logger.info(f"Returning single file: {filename}")
            
            return send_file(
                io.BytesIO(file_bytes),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            # Multiple sheets - return ZIP file
            logger.info(f"Creating ZIP with {len(split_files)} files")
            return create_zip_response(split_files, base_name)
            
    except RequestEntityTooLarge:
        logger.warning(f"File too large (max: {MAX_FILE_SIZE / (1024*1024):.1f} MB)")
        return jsonify({
            'error': f'File too large. Maximum size: {MAX_FILE_SIZE / (1024*1024):.1f} MB'
        }), 413
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred'}), 500


@app.errorhandler(413)
def request_entity_too_large(e):
    """Handle file size limit exceeded."""
    return jsonify({
        'error': f'File too large. Maximum size: {MAX_FILE_SIZE / (1024*1024):.1f} MB'
    }), 413


@app.errorhandler(500)
def internal_server_error(e):
    """Handle internal server errors."""
    logger.error(f"Internal server error: {str(e)}", exc_info=True)
    return jsonify({'error': 'Internal server error'}), 500


if __name__ == '__main__':
    logger.info(f"Starting Excel Splitter Service on port {PORT}")
    logger.info(f"Configuration: MAX_FILE_SIZE={MAX_FILE_SIZE_MB}MB, MAX_SHEETS={MAX_SHEETS}")
    app.run(host='0.0.0.0', port=PORT, debug=False)